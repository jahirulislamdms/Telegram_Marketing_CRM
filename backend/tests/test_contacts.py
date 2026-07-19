"""Phase 5 acceptance: contacts import (CSV+Excel), dedupe, resolution, messaging.

The engine is mocked (resolution/messaging need a live authorized account).
"""

import io

import pytest
from openpyxl import Workbook

from app.services import engine_client
from app.services.contacts import (
    normalize_phone,
    normalize_username,
    parse_bool,
    parse_csv,
)


# ------------------------------------------------------------------ pure -----


def test_normalize_phone():
    assert normalize_phone(" +1 (415) 555-0123 ") == "+14155550123"
    # A phone without '+' must gain one so it isn't read as a Telegram user id.
    assert normalize_phone("8801646562267") == "+8801646562267"
    assert normalize_phone("880 164 656 2267") == "+8801646562267"
    assert normalize_phone("") is None
    assert normalize_phone(None) is None
    assert normalize_phone("abc") is None


def test_normalize_username():
    assert normalize_username("@Sara_Ali") == "sara_ali"
    assert normalize_username("  Bob ") == "bob"
    assert normalize_username("") is None


def test_parse_bool():
    assert parse_bool("true") and parse_bool("1") and parse_bool("YES")
    assert not parse_bool("false")
    assert not parse_bool("")
    assert not parse_bool(None)


def test_parse_csv():
    data = b"name,phone,username,source,consent\nAhmed,+123,,offline,true\n"
    rows = parse_csv(data)
    assert rows[0]["name"] == "Ahmed"
    assert rows[0]["phone"] == "+123"
    assert rows[0]["consent"] == "true"


# ------------------------------------------------------------------ API ------


def _auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _login(client, email, password) -> str:
    r = client.post("/api/auth/login", json={"email": email, "password": password})
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


@pytest.fixture()
def admin_token(client, admin_credentials) -> str:
    return _login(client, **admin_credentials)


def _logged_in_account(client, token, monkeypatch) -> int:
    created = client.post(
        "/api/accounts",
        headers=_auth(token),
        json={"label": "resolver", "assign_proxy": False},
    )
    aid = created.json()["id"]

    async def _qr(_id):
        return {"status": "authorized", "url": None, "user": {"id": 7}, "detail": None}

    monkeypatch.setattr(engine_client, "qr_status", _qr)
    client.get(f"/api/accounts/{aid}/login/qr", headers=_auth(token))
    return aid


def test_import_csv_dedupe_and_consent(client, admin_token):
    csv_data = (
        "name,phone,username,source,consent\n"
        "Ahmed Khan,+923001234501,,offline_store,true\n"
        "Sara Ali,,@sara_ali_p5,online_store,true\n"
        ",+14155550501,,online_store,true\n"
        "Dup Phone,+923001234501,,x,true\n"
        "No Consent,+923001234502,,x,false\n"
        "Bad Row,,,x,true\n"
    ).encode()
    r = client.post(
        "/api/contacts/import",
        headers=_auth(admin_token),
        files={"file": ("contacts.csv", csv_data, "text/csv")},
    )
    assert r.status_code == 200, r.text
    body = r.json()
    # §15.3: a duplicate row now UPDATES the existing record instead of being skipped.
    assert body["imported"] == 3
    assert body["updated"] == 1  # "Dup Phone" matched Ahmed's +923001234501 and updated it
    assert body["skipped_duplicates"] == 0
    assert body["rejected_no_consent"] == 1
    assert body["invalid"] == 1
    assert body["errors"] == 0


def test_import_excel(client, admin_token):
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "phone", "username", "source", "consent"])
    ws.append(["Excel User", "", "@excel_user_p5", "online", "true"])
    ws.append(["", "+15550009901", "", "online", "true"])
    buf = io.BytesIO()
    wb.save(buf)
    r = client.post(
        "/api/contacts/import",
        headers=_auth(admin_token),
        files={
            "file": (
                "contacts.xlsx",
                buf.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 200, r.text
    assert r.json()["imported"] == 2


def test_import_template_download(client, admin_token):
    r = client.get("/api/contacts/import-template", headers=_auth(admin_token))
    assert r.status_code == 200
    assert "name,phone,username,source,consent" in r.text


def test_create_and_display_label(client, admin_token):
    r = client.post(
        "/api/contacts",
        headers=_auth(admin_token),
        json={"username": "@crud_user_p5", "consent": True, "source": "test"},
    )
    assert r.status_code == 201
    body = r.json()
    assert body["lead_type"] == "username"
    assert body["display_label"] == "@crud_user_p5"

    # Stage transitions.
    cid = body["id"]
    p = client.patch(
        f"/api/contacts/{cid}", headers=_auth(admin_token), json={"stage": "contacted"}
    )
    assert p.json()["stage"] == "contacted"
    p2 = client.patch(
        f"/api/contacts/{cid}", headers=_auth(admin_token), json={"stage": "opted_out"}
    )
    assert p2.json()["opted_out"] is True


def test_create_requires_identifier(client, admin_token):
    r = client.post(
        "/api/contacts",
        headers=_auth(admin_token),
        json={"name": "No Identifier", "consent": True},
    )
    assert r.status_code == 422


def test_resolve_username(client, admin_token, monkeypatch):
    _logged_in_account(client, admin_token, monkeypatch)
    contact = client.post(
        "/api/contacts",
        headers=_auth(admin_token),
        json={"username": "@resolvable_p5", "consent": True},
    ).json()

    async def _resolve(account, proxy, username):
        return {"user_id": 123456, "status": "resolved"}

    monkeypatch.setattr(engine_client, "resolve_username", _resolve)
    r = client.post(
        f"/api/contacts/{contact['id']}/resolve", headers=_auth(admin_token)
    )
    assert r.status_code == 200
    assert r.json()["resolution_status"] == "resolved"
    assert r.json()["telegram_user_id"] == 123456


def test_message_updates_stage(client, admin_token, monkeypatch):
    aid = _logged_in_account(client, admin_token, monkeypatch)
    contact = client.post(
        "/api/contacts",
        headers=_auth(admin_token),
        json={"username": "@msg_user_p5", "consent": True},
    ).json()

    sent = []

    async def _send(account, proxy, target, text):
        sent.append((target, text))
        return {"sent": True}

    monkeypatch.setattr(engine_client, "send_message", _send)
    r = client.post(
        f"/api/contacts/{contact['id']}/message",
        headers=_auth(admin_token),
        json={"account_id": aid, "text": "hello there"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["stage"] == "contacted"
    assert r.json()["last_contacted_at"] is not None
    assert sent and sent[0][1] == "hello there"


def test_send_identifier_prefers_resolvable_over_user_id():
    """A re-resolvable @username/+phone beats the cached numeric id, so ANY
    account can reach the contact (a user id's access-hash is per-account)."""
    from app.db.models.contact import Contact
    from app.services.contacts import send_identifier

    assert send_identifier(Contact(username="alice", phone="+123", telegram_user_id=9)) == "@alice"
    # A resolved phone lead still sends via +phone (imported per-account), NOT the id.
    assert (
        send_identifier(Contact(phone="+8801646562267", telegram_user_id=855963265073))
        == "+8801646562267"
    )
    assert send_identifier(Contact(phone="8801646562267")) == "+8801646562267"  # '+' added
    assert send_identifier(Contact(telegram_user_id=42)) == "42"  # last resort
    assert send_identifier(Contact()) is None


def test_message_phone_contact_sends_plus_prefixed_target(client, admin_token, monkeypatch):
    """A phone contact must reach the engine as '+phone' (so it resolves), not a
    bare number the engine would read as a user id."""
    aid = _logged_in_account(client, admin_token, monkeypatch)
    contact = client.post(
        "/api/contacts", headers=_auth(admin_token),
        json={"name": "Phone Lead", "phone": "8801646562267", "consent": True},
    ).json()
    # Stored normalised with a leading '+'.
    assert contact["phone"] == "+8801646562267"

    sent = []

    async def _send(account, proxy, target, text):
        sent.append(target)
        return {"sent": True}

    monkeypatch.setattr(engine_client, "send_message", _send)
    r = client.post(
        f"/api/contacts/{contact['id']}/message",
        headers=_auth(admin_token),
        json={"account_id": aid, "text": "hi"},
    )
    assert r.status_code == 200, r.text
    assert sent == ["+8801646562267"]  # phone with '+', engine will import/resolve it


def test_message_blocked_without_consent(client, admin_token, monkeypatch):
    aid = _logged_in_account(client, admin_token, monkeypatch)
    contact = client.post(
        "/api/contacts",
        headers=_auth(admin_token),
        json={"username": "@noconsent_p5", "consent": False},
    ).json()
    r = client.post(
        f"/api/contacts/{contact['id']}/message",
        headers=_auth(admin_token),
        json={"account_id": aid, "text": "hi"},
    )
    assert r.status_code == 403


def test_message_blocked_when_opted_out(client, admin_token, monkeypatch):
    aid = _logged_in_account(client, admin_token, monkeypatch)
    contact = client.post(
        "/api/contacts",
        headers=_auth(admin_token),
        json={"username": "@optout_p5", "consent": True},
    ).json()
    client.patch(
        f"/api/contacts/{contact['id']}",
        headers=_auth(admin_token),
        json={"stage": "opted_out"},
    )
    r = client.post(
        f"/api/contacts/{contact['id']}/message",
        headers=_auth(admin_token),
        json={"account_id": aid, "text": "hi"},
    )
    assert r.status_code == 403


def test_agent_sees_only_assigned(client, admin_token):
    client.post(
        "/api/users",
        headers=_auth(admin_token),
        json={"email": "agent5@test.com", "password": "AgentPass123", "role": "agent"},
    )
    agent_token = _login(client, "agent5@test.com", "AgentPass123")
    agent_id = client.get("/api/auth/me", headers=_auth(agent_token)).json()["id"]

    owned = client.post(
        "/api/contacts",
        headers=_auth(admin_token),
        json={"username": "@owned_p5", "consent": True},
    ).json()
    client.patch(
        f"/api/contacts/{owned['id']}",
        headers=_auth(admin_token),
        json={"assigned_agent_id": agent_id},
    )
    other = client.post(
        "/api/contacts",
        headers=_auth(admin_token),
        json={"username": "@notowned_p5", "consent": True},
    ).json()

    listing = client.get("/api/contacts", headers=_auth(agent_token)).json()
    ids = {c["id"] for c in listing}
    assert owned["id"] in ids
    assert other["id"] not in ids

    # Agents cannot import or view a contact that isn't theirs.
    assert (
        client.post(
            "/api/contacts/import",
            headers=_auth(agent_token),
            files={"file": ("c.csv", b"name,phone,username,source,consent\n", "text/csv")},
        ).status_code
        == 403
    )
    assert (
        client.get(f"/api/contacts/{other['id']}", headers=_auth(agent_token)).status_code
        == 403
    )


# ------------------------------------------------- §15.3 Contacts UX upgrade ---


def test_create_duplicate_phone_conflict(client, admin_token):
    payload = {"name": "Dupe A", "phone": "+9990000153001", "consent": True}
    r1 = client.post("/api/contacts", headers=_auth(admin_token), json=payload)
    assert r1.status_code == 201, r1.text
    r2 = client.post("/api/contacts", headers=_auth(admin_token), json=payload)
    assert r2.status_code == 409
    assert r2.json()["detail"] == "Phone number already exists."


def test_create_duplicate_username_conflict(client, admin_token):
    payload = {"username": "@dupe_user_153", "consent": True}
    assert client.post("/api/contacts", headers=_auth(admin_token), json=payload).status_code == 201
    r2 = client.post("/api/contacts", headers=_auth(admin_token), json=payload)
    assert r2.status_code == 409
    assert r2.json()["detail"] == "Username already exists."


def test_import_updates_existing_instead_of_duplicating(client, admin_token):
    first = (
        "name,phone,username,source,consent\n"
        "Imp One,+9990000153100,,src_a,true\n"
    ).encode()
    r1 = client.post(
        "/api/contacts/import",
        headers=_auth(admin_token),
        files={"file": ("c.csv", first, "text/csv")},
    )
    assert r1.json()["imported"] == 1
    # Re-import the same phone with a new name/source → updates, not a new row.
    again = (
        "name,phone,username,source,consent\n"
        "Imp One Renamed,+9990000153100,,src_b,true\n"
    ).encode()
    r2 = client.post(
        "/api/contacts/import",
        headers=_auth(admin_token),
        files={"file": ("c.csv", again, "text/csv")},
    )
    body = r2.json()
    assert body["imported"] == 0
    assert body["updated"] == 1
    # Confirm the single record was updated in place.
    listing = client.get(
        "/api/contacts?q=9990000153100", headers=_auth(admin_token)
    ).json()
    assert len(listing) == 1
    assert listing[0]["name"] == "Imp One Renamed"
    assert listing[0]["source"] == "src_b"


def test_edit_contact_fields(client, admin_token):
    created = client.post(
        "/api/contacts",
        headers=_auth(admin_token),
        json={"name": "Edit Me", "phone": "+9990000153200", "consent": True},
    ).json()
    r = client.patch(
        f"/api/contacts/{created['id']}",
        headers=_auth(admin_token),
        json={
            "name": "Edited Name",
            "username": "@edited_153",
            "source": "edited_src",
            "notes": "VIP lead — call first",
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["name"] == "Edited Name"
    assert body["username"] == "edited_153"  # normalised (lowercase, no '@')
    assert body["source"] == "edited_src"
    assert body["notes"] == "VIP lead — call first"
    assert body["phone"] == "+9990000153200"  # unchanged


def test_edit_into_duplicate_conflicts(client, admin_token):
    a = client.post(
        "/api/contacts", headers=_auth(admin_token),
        json={"username": "@edit_dupe_a153", "consent": True},
    ).json()
    client.post(
        "/api/contacts", headers=_auth(admin_token),
        json={"username": "@edit_dupe_b153", "consent": True},
    )
    # Renaming A's username to B's must 409.
    r = client.patch(
        f"/api/contacts/{a['id']}",
        headers=_auth(admin_token),
        json={"username": "@edit_dupe_b153"},
    )
    assert r.status_code == 409
    assert r.json()["detail"] == "Username already exists."


def test_bulk_consent_and_unresolve(client, admin_token):
    ids = []
    for i in range(2):
        ids.append(
            client.post(
                "/api/contacts", headers=_auth(admin_token),
                json={"username": f"@bulk153_{i}", "consent": True},
            ).json()["id"]
        )
    # Remove consent in bulk.
    r = client.post(
        "/api/contacts/bulk/consent",
        headers=_auth(admin_token),
        json={"contact_ids": ids, "consent": False},
    )
    assert r.status_code == 200 and r.json() == 2
    for cid in ids:
        assert client.get(f"/api/contacts/{cid}", headers=_auth(admin_token)).json()["consent"] is False
    # Unresolve in bulk resets resolution_status.
    r2 = client.post(
        "/api/contacts/bulk/unresolve",
        headers=_auth(admin_token),
        json={"contact_ids": ids},
    )
    assert r2.status_code == 200 and r2.json() == 2
    for cid in ids:
        got = client.get(f"/api/contacts/{cid}", headers=_auth(admin_token)).json()
        assert got["resolution_status"] == "pending"
        assert got["telegram_user_id"] is None


def test_export_csv_and_xlsx(client, admin_token):
    client.post(
        "/api/contacts", headers=_auth(admin_token),
        json={"name": "Export One", "phone": "+9990000153300", "source": "exp153", "consent": True},
    )
    # CSV, filtered to our unique source so other tests' rows don't interfere.
    rc = client.get("/api/contacts/export?format=csv&source=exp153", headers=_auth(admin_token))
    assert rc.status_code == 200
    assert rc.headers["content-type"].startswith("text/csv")
    text = rc.content.decode("utf-8-sig")
    assert "name,phone,username,source,stage,resolution,consent,created_at" in text
    assert "+9990000153300" in text
    # XLSX.
    rx = client.get("/api/contacts/export?format=xlsx&source=exp153", headers=_auth(admin_token))
    assert rx.status_code == 200
    assert "spreadsheetml" in rx.headers["content-type"]
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(rx.content))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0][0] == "name"
    assert any(r[1] == "+9990000153300" for r in rows[1:])


def test_list_pagination_and_total_header(client, admin_token):
    for i in range(5):
        client.post(
            "/api/contacts", headers=_auth(admin_token),
            json={"username": f"@page153_{i}", "source": "page153", "consent": True},
        )
    r = client.get(
        "/api/contacts?source=page153&limit=2&offset=0", headers=_auth(admin_token)
    )
    assert r.status_code == 200
    assert r.headers["X-Total-Count"] == "5"
    assert len(r.json()) == 2
    # Second page.
    r2 = client.get(
        "/api/contacts?source=page153&limit=2&offset=4", headers=_auth(admin_token)
    )
    assert len(r2.json()) == 1
