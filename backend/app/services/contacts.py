"""Contacts: CSV/Excel import, dedupe, CRUD, resolution, and messaging."""

import csv
import io
from datetime import datetime, timezone

from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models.account import Account
from app.db.models.contact import Contact
from app.db.models.proxy import Proxy
from app.services import engine_client

CSV_TEMPLATE = (
    "name,phone,username,source,consent\n"
    "Ahmed Khan,+923001234567,,offline_store,true\n"
    "Sara Ali,,@sara_ali,online_store,true\n"
    ",+14155550123,,online_store,true\n"
    ",,@no_name_user,online_store,true\n"
)

_TRUE_VALUES = {"true", "1", "yes", "y", "t"}


class NoResolverAccount(Exception):
    """No logged-in account is available to resolve/message a contact."""


class DuplicateContact(Exception):
    """A contact with the same phone or username already exists (§15.3).

    ``field`` is ``"phone"`` or ``"username"`` so the API can return the exact
    message the spec asks for ("Phone number already exists." / "Username already
    exists.").
    """

    def __init__(self, field: str) -> None:
        self.field = field
        label = "Phone number" if field == "phone" else "Username"
        self.message = f"{label} already exists."
        super().__init__(self.message)


# ------------------------------------------------------------- normalising ---


def normalize_phone(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return None
    # Always keep a leading '+' so a phone is never mistaken for a Telegram user
    # id (an all-digit target is otherwise sent as a numeric id and fails).
    return "+" + digits


def normalize_username(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip().lstrip("@").lower()
    return s or None


def parse_bool(value) -> bool:
    return str(value).strip().lower() in _TRUE_VALUES if value is not None else False


# ---------------------------------------------------------------- parsing ----


def _clean_headers(headers: list) -> list[str]:
    return [str(h).strip().lower() if h is not None else "" for h in headers]


def parse_csv(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for raw in reader:
        rows.append({(k or "").strip().lower(): v for k, v in raw.items()})
    return rows


def parse_xlsx(data: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = _clean_headers(list(next(rows_iter)))
    except StopIteration:
        return []
    rows = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        rows.append({headers[i]: values[i] if i < len(values) else None for i in range(len(headers))})
    wb.close()
    return rows


def parse_upload(filename: str, data: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return parse_xlsx(data)
    return parse_csv(data)


# ----------------------------------------------------------------- import ----


async def import_contacts(db: AsyncSession, rows: list[dict]) -> dict:
    """Import rows, **updating** any existing contact matched by phone/username
    instead of creating a duplicate (§15.3 #4/#5).

    Returns a summary: ``imported`` (new), ``updated`` (existing rows refreshed),
    ``rejected_no_consent``, ``invalid`` (no identifier), ``errors``, and ``total``.
    ``skipped_duplicates`` is retained (always 0 now) for backward compatibility.
    """
    result = await db.execute(select(Contact))
    by_phone: dict[str, Contact] = {}
    by_username: dict[str, Contact] = {}
    for c in result.scalars().all():
        if c.phone:
            by_phone[c.phone] = c
        if c.username:
            by_username[c.username] = c

    imported = updated = rejected = invalid = errors = 0

    for row in rows:
        try:
            name = (str(row.get("name")).strip() if row.get("name") else "") or None
            phone = normalize_phone(row.get("phone"))
            username = normalize_username(row.get("username"))
            source = (str(row.get("source")).strip() if row.get("source") else "") or None
            consent = parse_bool(row.get("consent"))

            if not phone and not username:
                invalid += 1
                continue
            if not consent:
                rejected += 1
                continue

            existing = None
            if phone and phone in by_phone:
                existing = by_phone[phone]
            elif username and username in by_username:
                existing = by_username[username]

            if existing is not None:
                # Update in place — never create a duplicate.
                if name:
                    existing.name = name
                if phone:
                    existing.phone = phone
                if username:
                    existing.username = username
                if source:
                    existing.source = source
                existing.consent = True
                existing.lead_type = "phone" if existing.phone else "username"
                if existing.phone:
                    by_phone[existing.phone] = existing
                if existing.username:
                    by_username[existing.username] = existing
                updated += 1
            else:
                contact = Contact(
                    name=name,
                    lead_type="phone" if phone else "username",
                    phone=phone,
                    username=username,
                    source=source,
                    consent=True,
                    tags=[],
                    utm={},
                )
                db.add(contact)
                if phone:
                    by_phone[phone] = contact
                if username:
                    by_username[username] = contact
                imported += 1
        except Exception:  # noqa: BLE001 — one bad row must not abort the batch
            errors += 1

    await db.commit()
    total = await db.scalar(select(func.count()).select_from(Contact))
    return {
        "imported": imported,
        "updated": updated,
        "skipped_duplicates": 0,
        "rejected_no_consent": rejected,
        "invalid": invalid,
        "errors": errors,
        "total": int(total or 0),
    }


# ------------------------------------------------------------------ CRUD -----


async def get_by_id(db: AsyncSession, contact_id: int) -> Contact | None:
    return await db.get(Contact, contact_id)


def _contacts_filter(
    stmt,
    *,
    assigned_agent_id: int | None,
    stage: str | None,
    source: str | None,
    resolution: str | None,
    lead_type: str | None,
    consent: bool | None,
    q: str | None,
    in_destination: int | None,
    not_in_destination: int | None,
):
    """Apply the shared Contacts filters to a SELECT (list + count share this)."""
    from app.db.models.destination import GroupMembership

    in_states = ("added", "invited", "joined")
    if assigned_agent_id is not None:
        stmt = stmt.where(Contact.assigned_agent_id == assigned_agent_id)
    if stage:
        stmt = stmt.where(Contact.stage == stage)
    if source:
        stmt = stmt.where(Contact.source == source)
    if resolution:
        stmt = stmt.where(Contact.resolution_status == resolution)
    if lead_type:
        stmt = stmt.where(Contact.lead_type == lead_type)
    if consent is not None:
        stmt = stmt.where(Contact.consent.is_(consent))
    if in_destination is not None:
        sub = select(GroupMembership.contact_id).where(
            GroupMembership.destination_id == in_destination,
            GroupMembership.state.in_(in_states),
        )
        stmt = stmt.where(Contact.id.in_(sub))
    if not_in_destination is not None:
        sub = select(GroupMembership.contact_id).where(
            GroupMembership.destination_id == not_in_destination,
            GroupMembership.state.in_(in_states),
        )
        stmt = stmt.where(Contact.id.notin_(sub))
    if q:
        like = f"%{q.lower()}%"
        stmt = stmt.where(
            or_(
                Contact.name.ilike(like),
                Contact.username.ilike(like),
                Contact.phone.ilike(like),
                Contact.source.ilike(like),
            )
        )
    return stmt


async def count_contacts(
    db: AsyncSession,
    *,
    assigned_agent_id: int | None = None,
    stage: str | None = None,
    source: str | None = None,
    resolution: str | None = None,
    lead_type: str | None = None,
    consent: bool | None = None,
    q: str | None = None,
    in_destination: int | None = None,
    not_in_destination: int | None = None,
) -> int:
    stmt = _contacts_filter(
        select(func.count()).select_from(Contact),
        assigned_agent_id=assigned_agent_id,
        stage=stage,
        source=source,
        resolution=resolution,
        lead_type=lead_type,
        consent=consent,
        q=q,
        in_destination=in_destination,
        not_in_destination=not_in_destination,
    )
    return int(await db.scalar(stmt) or 0)


async def list_contacts(
    db: AsyncSession,
    *,
    assigned_agent_id: int | None = None,
    stage: str | None = None,
    source: str | None = None,
    resolution: str | None = None,
    lead_type: str | None = None,
    consent: bool | None = None,
    q: str | None = None,
    in_destination: int | None = None,
    not_in_destination: int | None = None,
    limit: int | None = None,
    offset: int | None = None,
) -> list[Contact]:
    stmt = _contacts_filter(
        select(Contact),
        assigned_agent_id=assigned_agent_id,
        stage=stage,
        source=source,
        resolution=resolution,
        lead_type=lead_type,
        consent=consent,
        q=q,
        in_destination=in_destination,
        not_in_destination=not_in_destination,
    )
    stmt = stmt.order_by(Contact.created_at.desc(), Contact.id.desc())
    if offset:
        stmt = stmt.offset(offset)
    if limit is not None:
        stmt = stmt.limit(limit)
    result = await db.execute(stmt)
    return list(result.scalars().all())


async def find_conflict(
    db: AsyncSession,
    *,
    phone: str | None,
    username: str | None,
    exclude_id: int | None = None,
) -> str | None:
    """Return ``"phone"``/``"username"`` if another contact already uses it, else
    ``None`` (§15.3 #4 — phone & username are unique)."""
    if phone:
        stmt = select(Contact.id).where(Contact.phone == phone)
        if exclude_id is not None:
            stmt = stmt.where(Contact.id != exclude_id)
        if await db.scalar(stmt.limit(1)) is not None:
            return "phone"
    if username:
        stmt = select(Contact.id).where(Contact.username == username)
        if exclude_id is not None:
            stmt = stmt.where(Contact.id != exclude_id)
        if await db.scalar(stmt.limit(1)) is not None:
            return "username"
    return None


async def create_contact(
    db: AsyncSession,
    *,
    name: str | None,
    phone: str | None,
    username: str | None,
    source: str | None,
    consent: bool,
    tags: list | None,
    notes: str | None = None,
) -> Contact:
    phone = normalize_phone(phone)
    username = normalize_username(username)
    conflict = await find_conflict(db, phone=phone, username=username)
    if conflict:
        raise DuplicateContact(conflict)
    contact = Contact(
        name=name,
        lead_type="phone" if phone else "username",
        phone=phone,
        username=username,
        source=source,
        notes=notes,
        consent=consent,
        tags=tags or [],
        utm={},
    )
    db.add(contact)
    await db.commit()
    await db.refresh(contact)
    return contact


async def update_contact(db: AsyncSession, contact: Contact, **fields) -> Contact:
    for key, value in fields.items():
        if value is not None:
            setattr(contact, key, value)
    await db.commit()
    await db.refresh(contact)
    return contact


# Keys the Edit-contact modal may change (§15.3 #3). Others go through the
# generic PATCH path (stage/assignment) untouched.
_EDITABLE = {"name", "phone", "username", "source", "notes"}


async def edit_contact(db: AsyncSession, contact: Contact, updates: dict) -> Contact:
    """Apply an Edit-contact save: normalise phone/username, enforce uniqueness,
    and keep ``lead_type`` in sync. Unlike ``update_contact`` this **allows
    clearing** a field to ``None`` (e.g. removing a phone) as long as at least one
    identifier remains.
    """
    updates = dict(updates)
    if "phone" in updates:
        updates["phone"] = normalize_phone(updates["phone"])
    if "username" in updates:
        updates["username"] = normalize_username(updates["username"])

    new_phone = updates["phone"] if "phone" in updates else contact.phone
    new_username = updates["username"] if "username" in updates else contact.username
    if not new_phone and not new_username:
        raise ValueError("either phone or username is required")

    conflict = await find_conflict(
        db,
        phone=updates["phone"] if "phone" in updates else None,
        username=updates["username"] if "username" in updates else None,
        exclude_id=contact.id,
    )
    if conflict:
        raise DuplicateContact(conflict)

    for key in _EDITABLE:
        if key in updates:
            setattr(contact, key, updates[key])
    contact.lead_type = "phone" if contact.phone else "username"
    await db.commit()
    await db.refresh(contact)
    return contact


# ------------------------------------------------------------------ export ---

EXPORT_HEADERS = [
    "name",
    "phone",
    "username",
    "source",
    "stage",
    "resolution",
    "consent",
    "created_at",
]


def _export_row(c: Contact) -> list:
    return [
        c.name or "",
        c.phone or "",
        f"@{c.username}" if c.username else "",
        c.source or "",
        c.stage,
        c.resolution_status,
        "true" if c.consent else "false",
        c.created_at.isoformat() if c.created_at else "",
    ]


def contacts_to_csv(contacts: list[Contact]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(EXPORT_HEADERS)
    for c in contacts:
        writer.writerow(_export_row(c))
    return buf.getvalue().encode("utf-8-sig")


def contacts_to_xlsx(contacts: list[Contact]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(EXPORT_HEADERS)
    for c in contacts:
        ws.append(_export_row(c))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ------------------------------------------------------------ bulk helpers ---


async def bulk_set_consent(db: AsyncSession, contact_ids: list[int], consent: bool) -> int:
    count = 0
    for cid in contact_ids:
        contact = await db.get(Contact, cid)
        if contact is None:
            continue
        contact.consent = consent
        if consent:
            contact.opted_out = False
        count += 1
    await db.commit()
    return count


async def bulk_unresolve(db: AsyncSession, contact_ids: list[int]) -> int:
    count = 0
    for cid in contact_ids:
        contact = await db.get(Contact, cid)
        if contact is None:
            continue
        contact.resolution_status = "pending"
        contact.telegram_user_id = None
        count += 1
    await db.commit()
    return count


async def delete_contact(db: AsyncSession, contact: Contact) -> None:
    await db.delete(contact)
    await db.commit()


# ------------------------------------------------- resolution & messaging ----


async def _account_proxy(db: AsyncSession, account: Account) -> Proxy | None:
    if account.proxy_id is None:
        return None
    return await db.get(Proxy, account.proxy_id)


async def pick_resolver_account(db: AsyncSession, contact: Contact) -> Account | None:
    if contact.assigned_account_id:
        acc = await db.get(Account, contact.assigned_account_id)
        if acc and acc.session_ref and acc.status in ("active", "warming"):
            return acc
    result = await db.execute(
        select(Account)
        .where(
            Account.session_ref.isnot(None),
            Account.status.in_(["active", "warming"]),
        )
        .order_by(Account.id)
        .limit(1)
    )
    return result.scalar_one_or_none()


async def resolve_contact(db: AsyncSession, contact: Contact) -> Contact:
    account = await pick_resolver_account(db, contact)
    if account is None:
        raise NoResolverAccount()
    proxy = await _account_proxy(db, account)

    if contact.lead_type == "username" and contact.username:
        data = await engine_client.resolve_username(account, proxy, contact.username)
    elif contact.lead_type == "phone" and contact.phone:
        data = await engine_client.resolve_phone(account, proxy, contact.phone)
    else:
        contact.resolution_status = "failed"
        await db.commit()
        await db.refresh(contact)
        return contact

    contact.telegram_user_id = data.get("user_id")
    contact.resolution_status = data.get("status", "failed")
    await db.commit()
    await db.refresh(contact)
    return contact


def send_identifier(contact: Contact) -> str | None:
    """The best target to reach a contact from *any* account.

    Prefer a **re-resolvable** identifier (``@username`` / ``+phone``) over the
    cached numeric ``telegram_user_id``: a user id only carries an access-hash for
    the one account that resolved it, so sending it from a different account fails
    with "Could not find the input entity". A ``@username`` resolves publicly and a
    ``+phone`` is imported per-account, so every account can reach the contact.
    """
    if contact.username:
        return f"@{contact.username}"
    if contact.phone:
        # Keep the '+' so the engine treats it as a phone to import/resolve, not a
        # numeric user id (covers rows saved before the normalisation fix).
        return contact.phone if contact.phone.startswith("+") else f"+{contact.phone}"
    if contact.telegram_user_id:
        return str(contact.telegram_user_id)
    return None


# Backwards-compatible alias.
message_target = send_identifier


async def message_contact(
    db: AsyncSession, contact: Contact, account: Account, text: str
) -> Contact:
    proxy = await _account_proxy(db, account)
    target = message_target(contact)
    if target is None:
        raise ValueError("contact has no reachable identifier")
    result = await engine_client.send_message(account, proxy, target, text)
    if not result.get("sent", False):
        raise engine_client.EngineUnavailable(f"send failed: {result.get('error')}")
    contact.last_contacted_at = datetime.now(timezone.utc)
    if contact.stage == "new":
        contact.stage = "contacted"
    await db.commit()
    await db.refresh(contact)
    return contact
